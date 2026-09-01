import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

SRC = "outputs/live_coverage/coverage_audit.csv"
OUT = "outputs/live_coverage/asx_monitoring_funds.xlsx"
d = pd.read_csv(SRC)
a = d[(d["market_code"].astype(str).str.upper().isin(["ASX", "AU"]))
      & (d["monitoring_eligible"] == True)].copy()          # noqa: E712

# The audit ran before the NAV-continuity guard reached the live table, so it
# still calls Argo GREEN on the $5.00 anchor that is really $2.75 - the exact
# fabricated signal the guard exists to stop. Shipping the audit's status
# unmodified would put that back at the top of the green list. The live table
# is the newer authority on whether a row may alert, so it overrides here.
live = pd.read_parquet("data/nta_live/latest.parquet")
live["ticker"] = (live["security_id"].astype(str)
                  .str.replace("^ASX:", "", regex=True).str.upper())
keep = ["ticker", "nav_continuity_ok", "nav_prev", "nav_jump", "alert_eligible",
        "data_quality_reason"]
live = live[[c for c in keep if c in live.columns]].drop_duplicates("ticker")
a["ticker"] = a["ticker"].astype(str).str.upper()
a = a.merge(live, on="ticker", how="left", suffixes=("", "_live"))

held = ~a["nav_continuity_ok"].fillna(True).astype(bool)
a.loc[held, "coverage_status"] = "RED"
a.loc[held, "blocking_issue"] = "nav_discontinuity"
a.loc[held, "recommended_fix"] = (
    "the NAV we hold jumps implausibly from this fund's previous published "
    "figure - almost always the parser reading the wrong number off the page. "
    "Check the announcement before trusting any discount shown here.")

NAV_KIND = {
    "nta_announcement": "Fund's own NTA announcement",
    "previously_published_nta": "Fund's earlier published NTA",
    "asx_monthly_report": "ASX monthly report",
    "monthly_report": "Monthly report",
    "stale_historical_anchor": "Stale historical anchor",
}
# plain-English blocking reasons; the audit's own labels are accurate but terse
WHY = {
    "z_within_error_band": "Discount is not far enough from its own average to beat the uncertainty in our NAV",
    "rolled_forward_nav": "NAV is estimated forward from an older figure, not published",
    "insufficient_zscore_history": "Fewer than 24 months of discount history",
    "stale_nav": "Newest NAV we hold is old",
    "nav_too_stale": "NAV is past the age limit for alerting",
    "stale_panel_price_only": "No live price - only a month-end print",
    "suspected_unit_mismatch": "Price and NAV look to be in different units",
    "extreme_discount_premium": "Discount/premium is implausibly large - treated as a data fault",
    "no_nav": "No NAV from any source",
    "nav_not_positive": "NAV is zero or negative",
    "nav_discontinuity": ("NAV jumps implausibly from this fund's own previous figure "
                          "- treated as a bad reading, not a real move"),
}
a["Why not green"] = a["blocking_issue"].map(WHY)
# an unmapped label is shown verbatim rather than silently blanked - a blank
# would read as "nothing wrong", which is the opposite of the truth
a["Why not green"] = a["Why not green"].fillna(a["blocking_issue"]).fillna("")
a.loc[a["coverage_status"].eq("GREEN"), "Why not green"] = ""

order = {"GREEN": 0, "AMBER": 1, "RED": 2}
a["_o"] = a["coverage_status"].map(order).fillna(3)
a = a.sort_values(["_o", "blocking_issue", "ticker"])

cols = [
    ("ticker", "Ticker"), ("name", "Fund"), ("coverage_status", "Status"),
    ("price", "Price"), ("price_age_days", "Price age (d)"),
    ("nav", "NAV"), ("nav_effective_date", "NAV date"),
    ("nav_kind", "NAV source"), ("nav_staleness_days", "NAV age (d)"),
    ("nav_est_error", "NAV error"), ("discount", "Discount"),
    ("disc_mu_36m", "Avg discount 36m"), ("z_adj", "Z score"),
    ("nav_prev", "Previous NAV"), ("nav_jump", "NAV jump"),
    ("Why not green", "Why not green"), ("recommended_fix", "What would fix it"),
]
t = a[[c for c, _ in cols]].copy()
t["nav_kind"] = t["nav_kind"].map(NAV_KIND).fillna(t["nav_kind"])
t.columns = [h for _, h in cols]

wb = Workbook()
ws = wb.active
ws.title = "ASX funds"

TITLE = Font(name="Arial", size=14, bold=True)
HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)
NOTE = Font(name="Arial", size=9, italic=True, color="595959")
HFILL = PatternFill("solid", fgColor="1F4E4A")
G = PatternFill("solid", fgColor="D6EADF")
Y = PatternFill("solid", fgColor="FBF0D9")
R = PatternFill("solid", fgColor="F7DDD9")
thin = Side(style="thin", color="D9D9D9")
BORD = Border(bottom=thin)

ws["A1"] = "ASX monitoring-eligible funds"
ws["A1"].font = TITLE
ws["A2"] = (f"{len(t)} funds · as at {d['price_date'].dropna().max()} · "
            f"GREEN {int((t['Status']=='GREEN').sum())} · "
            f"AMBER {int((t['Status']=='AMBER').sum())} · "
            f"RED {int((t['Status']=='RED').sum())}")
ws["A2"].font = NOTE
ws["A3"] = ("GREEN = every input present and the signal clears the uncertainty in our own NAV. "
            "A blank cell means we hold no value - never a zero. Rows whose NAV jumps "
            "implausibly from the fund's own previous figure are forced to RED here, "
            "overriding the audit, which predates that check.")
ws["A3"].font = NOTE

HDR = 5
for j, h in enumerate(t.columns, start=1):
    c = ws.cell(row=HDR, column=j, value=h)
    c.font, c.fill = HEAD, HFILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i, (_, r) in enumerate(t.iterrows(), start=HDR + 1):
    for j, h in enumerate(t.columns, start=1):
        v = r[h]
        if pd.isna(v):
            v = None                      # missing stays missing, never 0
        c = ws.cell(row=i, column=j, value=v)
        c.font, c.border = BODY, BORD
        if h in ("Price", "NAV", "Avg discount 36m", "Discount", "NAV error",
                 "Z score", "Previous NAV", "NAV jump"):
            c.alignment = Alignment(horizontal="right")
        if h in ("Why not green", "What would fix it", "Fund"):
            c.alignment = Alignment(vertical="top", wrap_text=True)
    st = r["Status"]
    ws.cell(row=i, column=3).fill = {"GREEN": G, "AMBER": Y, "RED": R}.get(st, Y)
    ws.cell(row=i, column=3).alignment = Alignment(horizontal="center")

for col, fmt in (("D", "$#,##0.000"), ("F", "$#,##0.000"), ("J", "0.0%"),
                 ("K", "0.0%"), ("L", "0.0%"), ("M", "0.00"),
                 ("E", "0"), ("I", "0"), ("N", "$#,##0.000"), ("O", "0.0%")):
    for i in range(HDR + 1, HDR + 1 + len(t)):
        ws[f"{col}{i}"].number_format = fmt

widths = {"A": 9, "B": 34, "C": 9, "D": 10, "E": 11, "F": 10, "G": 12,
          "H": 26, "I": 10, "J": 10, "K": 10, "L": 15, "M": 9,
          "N": 12, "O": 10, "P": 46, "Q": 60}
for k, v in widths.items():
    ws.column_dimensions[k].width = v
ws.freeze_panes = f"A{HDR + 1}"
ws.auto_filter.ref = f"A{HDR}:{get_column_letter(len(t.columns))}{HDR + len(t)}"
ws.row_dimensions[HDR].height = 30

ws["J5"].comment = Comment(
    "Our estimated uncertainty in this fund's NAV. The Z score is voided when the "
    "discount's distance from its own average is smaller than this. Median by source: "
    "fund's own published NTA 1.9%, NTA announcement 6.2%, ASX monthly report 12.3%.",
    "coverage audit")
ws["M5"].comment = Comment(
    "Error-adjusted Z: how far today's discount sits from this fund's own 36-month "
    "average, after allowing for NAV uncertainty. Blank = could not be computed.",
    "coverage audit")

# ---- why-not-green summary, as formulas over the sheet above
s = wb.create_sheet("Why not green")
s["A1"] = "Why funds are not green"
s["A1"].font = TITLE
s["A2"] = "Counts recalculate from the 'ASX funds' tab."
s["A2"].font = NOTE
for j, h in enumerate(("Reason", "Funds", "% of all"), start=1):
    c = s.cell(row=4, column=j, value=h)
    c.font, c.fill = HEAD, HFILL
    c.alignment = Alignment(horizontal="center")
last = HDR + len(t)
seen = [x for x in t["Why not green"].value_counts().index if x]
rows = ["(green - nothing blocking)"] + seen
for i, why in enumerate(rows, start=5):
    s.cell(row=i, column=1, value=why).font = BODY
    crit = '""' if i == 5 else f'$A{i}'
    s.cell(row=i, column=2,
           value=f"=COUNTIF('ASX funds'!$P${HDR + 1}:$P${last},{crit})").font = BODY
    s.cell(row=i, column=3, value=f"=IFERROR(B{i}/$B${5 + len(rows)},0)").font = BODY
    s.cell(row=i, column=3).number_format = "0.0%"
tot = 5 + len(rows)
s.cell(row=tot, column=1, value="Total").font = Font(name="Arial", size=10, bold=True)
s.cell(row=tot, column=2, value=f"=SUM(B5:B{tot - 1})").font = Font(name="Arial", size=10, bold=True)
s.column_dimensions["A"].width = 62
s.column_dimensions["B"].width = 10
s.column_dimensions["C"].width = 10
s[f"A{tot + 2}"] = ("Source: outputs/live_coverage/coverage_audit.csv, generated by the "
                    "live coverage audit. A fund may have several problems; the one shown "
                    "is the audit's primary blocking issue.")
s[f"A{tot + 2}"].font = NOTE

wb.save(OUT)
print("wrote", OUT, "| funds:", len(t))
