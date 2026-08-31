"""On-demand coverage audit: for every live fund we intend to monitor, do
we actually hold a usable price, NAV, discount and signal - and if not, why?

    python -m cef_live.coverage_audit              # audit what is on disk
    python -m cef_live.coverage_audit --refresh    # refresh first, then audit

MANUAL ONLY. Nothing here is scheduled: no cron, no nightly hook. It is a
diagnostic you run when you want to know where coverage stands, and the
GitHub workflow that runs it is `workflow_dispatch`-only by design.

What it is, and is not
----------------------
It is a READ-ACROSS of the existing system: the registry (universe.py), the
evidence-based liveness classifier (liveness.py), the research-policy
exclusions (eligibility.py), the live NTA table (nta_live.py), the NAV
harvesters' own output, and the committed announcement archives. It builds
no new store, fetches nothing on its own, and re-implements no parser.

It is not a fixer. Where a number looks wrong it says so - `unit_check_status`,
`blocking_issue`, `recommended_fix` - and leaves the number alone. The one
exception is deterministic and named: a UK NAV anchor that the live table
took through the old pounds path is RE-DERIVED from the same committed
source through the canonical (pence) reader, with the stored value kept
beside it in `nav_stored_raw`. That is re-reading a source correctly, not
adjusting an output.

Denominators
------------
Five, reported separately, because the interesting question is usually
which one a number is over:

  registry_total        every vehicle the registry has ever listed
  registry_live         what the AIC/ASX file still lists (aggregator view)
  liveness_live         what the funds' own filings say is alive
  research_eligible     what the research policy would hold (no VCTs,
                        no split-capital or ZDP lines, sterling UK quotes)
  monitoring_eligible   alive AND research-eligible - the audit's denominator

Excluded vehicles keep a row in the Excluded tab with the reason. Nothing
is deleted to improve a percentage.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
import yaml

from . import eligibility, liveness, units

OUT_DIR = Path("outputs/live_coverage")

GREEN, AMBER, RED, EXCLUDED = "GREEN", "AMBER", "RED", "EXCLUDED"

DEFAULT_THRESHOLDS = {
    "price": {"fresh_days": 5, "stale_days": 30},
    "nav": {"fresh_days": 35, "amber_days": 120},
    "zscore": {"min_months": 24},
}

BASIS_LABEL = {
    0: "0 - published NAV (issuer announcement)",
    1: "1 - factor roll-forward from last published NAV",
    2: "2 - holdings-based",
    3: "3 - stale (last published value carried)",
}

# what to do about each blocking issue, so the table is a work list
FIXES = {
    "ticker_unresolved":
        "resolve the TIDM/ASX code (python -m cef_live.cli resolve-tickers) "
        "or add it to config/investegate_tickers.csv once verified",
    "no_price":
        "check the fund still trades under this ticker; re-run the price "
        "layer for it and record the provider error",
    "stale_panel_price_only":
        "the live price feed returned nothing for this symbol - verify the "
        "Yahoo symbol and re-run; never present the panel price as today's",
    "stale_price":
        "re-run the price layer; if it keeps returning an old close the "
        "line may be suspended or the symbol wrong",
    "no_nav":
        "no NAV from any source - confirm the fund publishes one, then add "
        "its announcement route (Tier 0 harvest / ASX index)",
    "nav_announcement_unparsed":
        "a NAV announcement exists but the parser produced no value - add a "
        "rule for its layout in harvest_nav.UK_RULES / the ASX parser",
    "nav_too_stale":
        "the newest NAV we hold is older than the amber window - re-run the "
        "NAV harvest for this fund and check its publication frequency",
    "suspected_unit_mismatch":
        "do NOT rescale the output - find which side is wrong (NAV parse or "
        "quote currency) and fix the source reader",
    "extreme_discount_premium":
        "verify price and NAV against the fund's own latest announcement "
        "before trusting the discount",
    "insufficient_zscore_history":
        "the fund needs its own discount history in the monthly panel; "
        "until then it can be monitored but not z-scored",
    "rolled_forward_nav":
        "none needed - the NAV is modelled forward from a published anchor "
        "and carries its own error band",
    "monthly_report_nav_not_announcement":
        "the NTA comes from the ASX monthly investment-products report - a "
        "legitimate but monthly source. Harvest the fund's own NTA "
        "announcement (Tier 0) to get NAV at announcement frequency",
    "aggregator_nav_only":
        "the fund's own NAV route is missing - resolve its ticker and add it "
        "to the Tier 0 harvest so it stops depending on the aggregator",
    "stale_nav":
        "re-run the NAV harvest; the anchor is past the fresh window",
    "future_dated_price":
        "the price carries a date later than the audit date - check the "
        "provider's timezone handling before trusting it",
    "future_dated_nav":
        "the NAV carries a date later than the audit date - check the "
        "as-at date parsed from the announcement",
    "not_in_live_table":
        "no row in data/nta_live/latest.parquet: the table was built when a "
        "fund with no NAV anchor was dropped outright, which discarded its "
        "price too. nta_live now keeps the row - re-run the nightly (or this "
        "audit with --refresh) and the price, if there is one, will appear",
}


# ------------------------------------------------------------------ inputs

def load_params(path: str = "config/params.yaml") -> dict:
    raw = Path(path).read_text()
    raw = re.sub(r"\$\{[^}]+\}", "", raw)
    return yaml.safe_load(raw) or {}


def thresholds(params: dict) -> dict:
    got = (params or {}).get("coverage_audit", {}) or {}
    out = {k: dict(v) for k, v in DEFAULT_THRESHOLDS.items()}
    for k, v in got.items():
        if k in out and isinstance(v, dict):
            out[k].update(v)
    return out


def load_registry(params: dict, as_of: date | None = None,
                  registry_path: str = "data/universe/registry.parquet",
                  evidence: pd.DataFrame | None = None) -> pd.DataFrame:
    """Registry + evidence-based liveness + research-policy eligibility.

    The saved `status` is NOT trusted on its own: liveness is recomputed
    from the funds' own filings every time this runs, so the audit reports
    the best live-status logic the repo has rather than whatever a file
    last happened to hold. Both verdicts are kept side by side.
    """
    reg = pd.read_parquet(registry_path)
    if evidence is None:
        from .cli import _liveness_evidence
        evidence = _liveness_evidence()
    reg = liveness.apply(reg, evidence, as_of=as_of, params=params)
    reg = eligibility.classify(reg, liveness.LIVE_STATUSES)
    return reg


def load_live_table(path: str = "data/nta_live/latest.parquet") -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        return pd.DataFrame()
    return pd.read_parquet(p)


def load_tickers() -> pd.DataFrame:
    """security_id -> ticker, from the resolver's cache plus manual verifications."""
    frames = []
    p = Path("config/resolved_tickers.csv")
    if p.exists():
        t = pd.read_csv(p)
        t = t[t.get("status").eq("verified")] if "status" in t.columns else t
        frames.append(t.assign(ticker_status="verified")[
            ["security_id", "ticker", "ticker_status"]])
    p = Path("config/investegate_tickers.csv")
    if p.exists():
        t = pd.read_csv(p, comment="#")
        if {"security_id", "ticker"} <= set(t.columns):
            frames.append(t.assign(ticker_status="verified_manual")[
                ["security_id", "ticker", "ticker_status"]])
    if not frames:
        return pd.DataFrame(columns=["security_id", "ticker", "ticker_status"])
    out = pd.concat(frames, ignore_index=True)
    return out.dropna(subset=["ticker"]).drop_duplicates("security_id", keep="first")


def load_tier0(market: str) -> pd.DataFrame:
    """The harvesters' own most recent published NAVs, as committed."""
    f = Path("data/nta_live/uk_tier0_latest.csv" if market == "UK"
             else "data/nta_live/au_tier0_latest.csv")
    cols = ["security_id", "nav_date", "nav_value", "unit", "source", "headline"]
    if not f.exists():
        return pd.DataFrame(columns=cols)
    df = pd.read_csv(f)
    for c in cols:
        if c not in df.columns:
            df[c] = None
    df["nav_date"] = pd.to_datetime(df["nav_date"], errors="coerce")
    return df.dropna(subset=["nav_date"])[cols]


def uk_nav_archive_facts() -> pd.DataFrame:
    """Per UK fund: what its NAV RNS archive shows, parsed and unparsed.

    Every row of data/uk_nav_history*.parquet is a "Net Asset Value(s)"
    announcement (scripts/archive_uk_navs.py filters on that headline), so
    the archive answers two different questions that must not be merged:
    when did the fund last PUBLISH a NAV, and when did we last PARSE one.

    The shards overlap - a re-parse run (archive_uk_navs.py --reparse)
    writes a second row for an announcement the original crawl already
    recorded - so rows are deduplicated on ann_id with the PARSED outcome
    winning. Counting both would report a fund's announcements twice and
    would leave a superseded failure standing next to its own fix.
    """
    tick = load_tickers()
    by_tk = {str(r.ticker).upper(): r.security_id
             for r in tick.itertuples(index=False)}

    frames = []
    # the daily NAV panel's rows are, by construction, announcements a value
    # WAS extracted from - including ones a re-parse recovered after the
    # original crawl recorded them unparsed
    try:
        from . import uk_nav_panel as UKP
        panel = UKP.read_panel()
    except Exception:  # noqa: BLE001
        panel = None
    if panel is not None and len(panel):
        frames.append(pd.DataFrame({
            "ticker": panel["ticker"],
            "ann_id": panel["ann_id"],
            "ann_date": panel["published_at"],
            "status": "parsed",
        }))
    for f in sorted(Path("data").glob("uk_nav_history*.parquet")):
        try:
            h = pd.read_parquet(f)
        except Exception:  # noqa: BLE001
            continue
        if not {"ticker", "ann_date"} <= set(h.columns):
            continue
        keep = h[[c for c in ("ticker", "ann_id", "ann_date", "status")
                  if c in h.columns]].copy()
        if "ann_id" not in keep.columns:
            keep["ann_id"] = None
        if "status" not in keep.columns:
            keep["status"] = "parsed"
        frames.append(keep)
    if not frames:
        return pd.DataFrame(columns=["security_id", "nav_ann_count",
                                     "nav_ann_parsed", "last_nav_announcement",
                                     "last_parsed_nav_date", "last_ann_status",
                                     "nav_parse_rate"])

    h = pd.concat(frames, ignore_index=True)
    h["ticker"] = h["ticker"].astype(str).str.upper()
    h["d"] = pd.to_datetime(h["ann_date"], errors="coerce")
    h = h.dropna(subset=["d"])
    h["security_id"] = h["ticker"].map(by_tk)
    h = h.dropna(subset=["security_id"])
    # a re-parse supersedes the original attempt for the same announcement
    h["_ok"] = (h["status"] == "parsed").astype(int)
    keyed = h[h["ann_id"].notna()].sort_values("_ok").drop_duplicates(
        "ann_id", keep="last")
    unkeyed = h[h["ann_id"].isna()]
    h = pd.concat([keyed, unkeyed], ignore_index=True)

    rows = []
    for sid, g in h.groupby("security_id"):
        parsed = g[g["status"] == "parsed"]
        newest = g.sort_values("d").iloc[-1]
        rows.append({
            "security_id": sid,
            "nav_ann_count": int(len(g)),
            "nav_ann_parsed": int(len(parsed)),
            "last_nav_announcement": newest["d"].date().isoformat(),
            "last_parsed_nav_date": (parsed["d"].max().date().isoformat()
                                     if len(parsed) else None),
            "last_ann_status": str(newest["status"]),
            "nav_parse_rate": round(len(parsed) / len(g), 4) if len(g) else None,
        })
    return pd.DataFrame(rows)


ASX_NTA = re.compile(r"net tangible asset|\bNTA\b|net asset value|\bNAV\b", re.I)
ASX_MONTHLY = re.compile(r"monthly (?:report|update|nta|investment)|"
                         r"fund update|investment update", re.I)


def asx_nav_announcement_facts(
        index_path: str = "data/asx_ann_cache/asx1/lic_announcement_index.parquet"
) -> pd.DataFrame:
    """Per ASX code: the most recent NTA statement, monthly report, or any filing.

    The three are distinguished because they are different sources of an
    NTA: a "Net Tangible Asset Backing" release is the fund stating its NTA
    directly; a monthly report carries one inside a longer document; any
    other filing carries none at all.
    """
    p = Path(index_path)
    cols = ["security_id", "last_nta_announcement", "last_nta_headline",
            "last_nta_kind", "last_monthly_report", "last_any_announcement",
            "nta_announcements_90d"]
    if not p.exists():
        return pd.DataFrame(columns=cols)
    idx = pd.read_parquet(p)
    idx["d"] = pd.to_datetime(idx["release_date"], utc=True, errors="coerce")
    idx = idx.dropna(subset=["d"])
    idx["head"] = idx["headline"].fillna("")
    cutoff = pd.Timestamp.now(tz="UTC") - pd.Timedelta(days=90)
    rows = []
    for code, g in idx.groupby("code"):
        nta = g[g["head"].str.contains(ASX_NTA)]
        monthly = g[g["head"].str.contains(ASX_MONTHLY)]
        top = nta.sort_values("d").iloc[-1] if len(nta) else None
        kind = None
        if top is not None:
            kind = ("monthly_report" if ASX_MONTHLY.search(top["head"])
                    else "nta_announcement")
        rows.append({
            "security_id": f"ASX:{str(code).upper()}",
            "last_nta_announcement": (top["d"].date().isoformat()
                                      if top is not None else None),
            "last_nta_headline": (str(top["head"])[:120]
                                  if top is not None else None),
            "last_nta_kind": kind,
            "last_monthly_report": (monthly["d"].max().date().isoformat()
                                    if len(monthly) else None),
            "last_any_announcement": g["d"].max().date().isoformat(),
            "nta_announcements_90d": int((nta["d"] >= cutoff).sum()) if len(nta) else 0,
        })
    return pd.DataFrame(rows, columns=cols)


def own_nav_latest(market: str) -> pd.DataFrame:
    """Latest NAV per fund from OUR extraction, in the canonical unit.

    Read through cli._own_nav_history, which is the same reader the nightly
    uses - so a unit fix there is a unit fix here, and the audit cannot
    drift away from the pipeline it is auditing.
    """
    from .cli import _own_nav_history
    try:
        own = _own_nav_history(market)
    except Exception as exc:  # noqa: BLE001
        print(f"own NAV history unavailable for {market}: {exc}")
        own = pd.DataFrame()
    if own is None or not len(own):
        return pd.DataFrame(columns=["security_id", "own_nav_date",
                                     "own_nav_value", "own_nav_unit"])
    own = own.copy()
    own["nav_date"] = pd.to_datetime(own["nav_date"], errors="coerce")
    own = own.dropna(subset=["nav_date", "nav_value"])
    own = own.sort_values("nav_date").groupby("security_id").tail(1)
    unit = own["nav_unit"] if "nav_unit" in own.columns else units.CANONICAL_UNIT.get(market)
    return pd.DataFrame({
        "security_id": own["security_id"].values,
        "own_nav_date": own["nav_date"].dt.date.astype(str).values,
        "own_nav_value": own["nav_value"].astype(float).values,
        "own_nav_unit": unit if isinstance(unit, str) else unit.values,
    })


# ------------------------------------------------------------------ helpers

def _age_days(d, today: date):
    if d is None or (isinstance(d, float) and pd.isna(d)) or d == "":
        return None
    t = pd.to_datetime(d, errors="coerce", utc=True)
    if pd.isna(t):
        return None
    return int((pd.Timestamp(today, tz="UTC") - t.normalize()).days)


def _trading_age(d, today: date):
    if d is None or (isinstance(d, float) and pd.isna(d)) or d == "":
        return None
    t = pd.to_datetime(d, errors="coerce")
    if pd.isna(t):
        return None
    return int(np.busday_count(t.date(), today)) if t.date() <= today else 0


def _parse_price_asof(value: str) -> tuple[str | None, str | None]:
    """('yahoo:CTY.L', '2026-08-27') from the live table's packed field."""
    if not isinstance(value, str) or not value or value == "none":
        return None, None
    if value == "aggregator_panel":
        return "aggregator_panel", None
    if "@" in value:
        src, _, when = value.partition("@")
        return src, when
    return value, None


def _sedol(security_id: str) -> str | None:
    s = str(security_id)
    return s.split(":", 1)[1] if s.startswith("SEDOL:") else None


def _asx_code(security_id: str) -> str | None:
    s = str(security_id)
    return s.split(":", 1)[1] if s.startswith("ASX:") else None


# ------------------------------------------------------------------ the audit

def build_rows(reg: pd.DataFrame, live: pd.DataFrame, params: dict,
               today: date | None = None) -> pd.DataFrame:
    """One row per registry vehicle, with every coverage fact attached.

    Excluded vehicles are kept (with their reason) so the denominators can
    be reconciled from the row-level data rather than taken on trust.
    """
    today = today or datetime.now(timezone.utc).date()
    thr = thresholds(params)
    tick = load_tickers().set_index("security_id")
    asx_ann = asx_nav_announcement_facts().set_index("security_id") \
        if len(asx_nav_announcement_facts()) else pd.DataFrame()
    uk_ann = uk_nav_archive_facts()
    uk_ann = uk_ann.set_index("security_id") if len(uk_ann) else pd.DataFrame()
    tier0 = {m: load_tier0(m).sort_values("nav_date").groupby("security_id").tail(1)
             .set_index("security_id") for m in ("UK", "AU")}
    own = {}
    for m in ("UK", "AU"):
        o = own_nav_latest(m)
        own[m] = o.set_index("security_id") if len(o) else pd.DataFrame()

    live_ix = live.set_index("security_id") if len(live) else pd.DataFrame()
    zmin = int(thr["zscore"]["min_months"])

    rows = []
    for r in reg.to_dict("records"):
        sid = str(r["security_id"])
        market = str(r.get("market", ""))
        lv = live_ix.loc[sid].to_dict() if sid in getattr(live_ix, "index", []) else {}
        if isinstance(lv.get("market"), pd.Series):        # duplicate ids
            lv = {}
        tk = tick.loc[sid].to_dict() if sid in getattr(tick, "index", []) else {}

        # ---------------- price
        price_stored = lv.get("price")
        src, when = _parse_price_asof(lv.get("price_asof"))
        price_source = lv.get("price_source") or src
        price_date = lv.get("price_date") or when
        fallback = lv.get("price_is_fallback")
        if fallback is None:
            fallback = (price_source == "aggregator_panel")
        fallback = bool(fallback) and price_stored is not None and pd.notna(price_stored)
        price_ccy = lv.get("price_ccy")
        # a monthly panel print is dated to a month, not a day
        if fallback and price_date is None:
            anchor = lv.get("price_panel_month") or (
                src if src and re.fullmatch(r"\d{4}-\d{2}", str(src)) else None)
            price_date = anchor
        price_age = _age_days(price_date, today)
        price_value, price_unit, price_unit_note = units.normalise(
            market, price_stored, price_ccy)
        ticker = tk.get("ticker")
        ticker_status = tk.get("ticker_status")
        if market == "AU" and not ticker:
            # an ASX security IS its code: the price layer builds `{code}.AX`
            # directly, so there is nothing to resolve and calling it
            # "unresolved" invented 107 ticker problems that do not exist
            ticker = _asx_code(sid)
            ticker_status = "asx_code" if ticker else "unresolved"
        if not ticker_status:
            ticker_status = "unresolved" if not ticker else "verified"
        price_error = None
        if price_stored is None or pd.isna(price_stored):
            # "we never asked" and "we asked and got nothing" are different
            # failures with different fixes
            price_error = (
                "fund_absent_from_live_table_never_priced" if not lv
                else "no_price_returned_for_symbol" if ticker
                else "no_ticker_to_price_with")
        # a fallback price is NEVER a current market price, whatever its age
        price_fresh = bool(
            not fallback and price_age is not None
            and price_age <= int(thr["price"]["fresh_days"]))

        # ---------------- NAV
        nav_stored = lv.get("nav_anchor")
        nav_date = lv.get("anchor_date")
        nav_source = lv.get("anchor_source")
        basis = lv.get("basis")
        nav_unit_lbl = lv.get("nav_unit") or units.CANONICAL_UNIT.get(market, "")
        nav_value = nav_stored
        nav_rederived = "no"

        o = own[market].loc[sid].to_dict() \
            if len(own.get(market, [])) and sid in own[market].index else {}
        t0 = tier0[market].loc[sid].to_dict() \
            if len(tier0.get(market, [])) and sid in tier0[market].index else {}

        # The one deterministic correction this module makes: a UK anchor the
        # live table took through the old pounds path is re-read from the same
        # committed archive through the canonical (pence) reader. The stored
        # value stays in nav_stored_raw.
        if (isinstance(nav_source, str) and nav_source.startswith("own_nav_history")
                and o.get("own_nav_value") is not None):
            canon, unit_lbl, _ = units.normalise(
                market, o["own_nav_value"], o.get("own_nav_unit"))
            if canon is not None and (nav_stored is None or pd.isna(nav_stored)
                                      or abs(canon - float(nav_stored)) > 1e-9):
                nav_value, nav_unit_lbl = canon, unit_lbl
                nav_rederived = "yes_canonical_unit_from_source"
            nav_date = o.get("own_nav_date") or nav_date

        # a Tier 0 announcement newer than whatever the table holds
        if t0.get("nav_value") is not None and pd.notna(t0.get("nav_value")):
            t0_date = pd.Timestamp(t0["nav_date"]).date().isoformat()
            if nav_date is None or str(t0_date) >= str(nav_date):
                canon, unit_lbl, _ = units.normalise(
                    market, t0["nav_value"], t0.get("unit"))
                nav_value, nav_unit_lbl = canon, unit_lbl
                nav_date, nav_source, basis = t0_date, str(t0.get("source")), 0
                if nav_rederived == "no":
                    nav_rederived = "tier0_announcement"

        nav_age = _age_days(nav_date, today)
        nav_published = bool(basis == 0 or (isinstance(nav_source, str) and (
            nav_source.startswith("own_nav_history")
            or nav_source.startswith("investegate")
            or nav_source.startswith("asx_ann"))))
        if market == "UK":
            nav_kind = ("newly_published_nav_announcement" if basis == 0 else
                        "previously_published_nav" if nav_published else
                        "rolled_forward_model" if basis == 1 else
                        "stale_historical_anchor")
        else:
            # The AU panel is the ASX's OWN monthly investment-products
            # report - an exchange publication of each LIC's NTA, not a
            # third-party aggregator's editorial pick. It is a legitimate
            # NTA source, just a monthly-frequency one, and the brief asks
            # for it to be distinguished from an NTA announcement rather
            # than lumped in with "aggregator only".
            k = (asx_ann.loc[sid].get("last_nta_kind")
                 if len(asx_ann) and sid in asx_ann.index else None)
            from_panel = isinstance(nav_source, str) and \
                nav_source.startswith("aggregator_panel")
            nav_kind = ("nta_announcement" if basis == 0 and k != "monthly_report"
                        else "monthly_report" if k == "monthly_report" and basis == 0
                        else "previously_published_nta" if nav_published else
                        "asx_monthly_report" if from_panel else
                        "rolled_forward_model" if basis == 1 else
                        "stale_historical_anchor")

        # NAV announcement vs NAV value: two different facts
        if market == "UK":
            a = uk_ann.loc[sid].to_dict() if len(uk_ann) and sid in uk_ann.index else {}
            last_ann = a.get("last_nav_announcement")
            last_parsed = a.get("last_parsed_nav_date")
            parse_rate = a.get("nav_parse_rate")
            ann_count = a.get("nav_ann_count")
        else:
            a = asx_ann.loc[sid].to_dict() if len(asx_ann) and sid in asx_ann.index else {}
            last_ann = a.get("last_nta_announcement")
            a = {**a, "last_ann_status": None}     # ASX index records no status
            # "when did we last get a VALUE out of one of this fund's
            # announcements" is the deterministic extract as much as it is
            # Tier 0. Counting only Tier 0 reported 46 ASX funds as
            # "announcement held, never parsed" while the extractor had in
            # fact read them - a parser failure that was ours to claim, not
            # theirs, and one that would have sent someone to fix a parser
            # that already worked.
            last_parsed = max(
                [x for x in (nav_date if basis == 0 else None,
                             o.get("own_nav_date"),
                             (pd.Timestamp(t0["nav_date"]).date().isoformat()
                              if t0.get("nav_date") is not None else None))
                 if x], default=None)
            parse_rate = None
            ann_count = a.get("nta_announcements_90d")
        unparsed = bool(last_ann and (not last_parsed or str(last_parsed) < str(last_ann)))
        # "we tried and no value came out" is a PARSER failure; "we hold the
        # announcement but never extracted from it" is a harvest gap. Only
        # the UK archive records which, because it stores a per-announcement
        # status; for ASX the neutral label is used rather than a guess.
        recorded_fail = bool(a.get("last_ann_status")
                             and str(a.get("last_ann_status")) != "parsed")
        if last_ann is None:
            parser_status = "no_nav_announcement_held"
        elif not unparsed:
            parser_status = "parsed"
        elif recorded_fail:
            parser_status = "parse_failed_recorded"
        elif not last_parsed:
            parser_status = "announcement_held_never_parsed"
        else:
            parser_status = "announcement_newer_than_last_parse"
        parse_reason = None
        if parser_status == "parse_failed_recorded":
            parse_reason = (f"the parser ran on the {last_ann} announcement and "
                            "produced no value - no rule matched this layout")
        elif parser_status == "announcement_held_never_parsed":
            parse_reason = ("a NAV/NTA announcement is indexed but no value has "
                            "been extracted from it")
        elif parser_status == "announcement_newer_than_last_parse":
            parse_reason = (f"latest NAV announcement {last_ann} not parsed; "
                            f"newest parsed value is {last_parsed}")

        # ---------------- units + discount
        diag = units.scale_diagnosis(price_value, nav_value)
        ccy_conflict = units.unit_metadata_conflict(market, price_ccy)
        if ccy_conflict and diag["unit_check_status"] == "ok":
            diag["unit_check_status"] = "metadata_conflict"
        if ccy_conflict:
            diag["unit_check_reason"] = "; ".join(
                x for x in (diag["unit_check_reason"], ccy_conflict) if x)
        disc = units.discount(price_value, nav_value)

        # price against the fund's OWN market panel print - never another
        # market's, and explicitly unavailable when the panel is not on disk
        panel_price = lv.get("price_panel")
        if panel_price is not None and pd.notna(panel_price) and panel_price > 0 \
                and price_value is not None and price_value > 0:
            pv = float(price_value) / float(panel_price)
            price_hist_check = ("ok" if 0.2 <= pv <= 5.0 else "implausible_vs_panel")
        else:
            pv = None
            price_hist_check = ("no_panel_price_held" if not fallback
                                else "price_is_the_panel_price")

        mu, sd = lv.get("disc_mu_36m"), lv.get("disc_sigma_36m")
        zscore_ok = bool(sd is not None and pd.notna(sd) and float(sd) > 0)

        rows.append({
            # identity
            "market": "ASX" if market == "AU" else market,
            "market_code": market,
            "ticker": ticker,
            "name": r.get("name"),
            "security_id": sid,
            "isin": r.get("isin"),
            "sedol": _sedol(sid),
            "asx_code": _asx_code(sid),
            "sector": r.get("sector"),
            "share_type": r.get("share_type"),
            "research_eligible": bool(r.get("research_eligible", False)),
            "monitoring_eligible": bool(r.get("monitoring_eligible", False)),
            "live_status": r.get("status"),
            "live_status_reason": r.get("liveness_reason"),
            "live_status_source": r.get("live_status_source"),
            "aggregator_status": r.get("aggregator_status"),
            "exclusion_reason": r.get("exclusion_reason"),
            # price
            "price": price_value,
            "price_unit": price_unit,
            "price_raw": price_stored,
            "price_raw_ccy": price_ccy,
            "price_unit_note": price_unit_note,
            "price_date": price_date,
            "price_source": price_source,
            "price_age_days": price_age,
            "price_age_trading_days": _trading_age(price_date, today)
                                      if price_date and len(str(price_date)) >= 10 else None,
            "price_is_fresh": price_fresh,
            "price_is_fallback_panel": fallback,
            "ticker_status": ticker_status,
            "price_error": price_error,
            # NAV
            "nav": nav_value,
            "nav_unit": nav_unit_lbl,
            "nav_stored_raw": nav_stored,
            "nav_effective_date": nav_date,
            "nav_announcement_date": last_ann,
            "nav_source": nav_source,
            "nav_basis": basis,
            "nav_basis_label": BASIS_LABEL.get(basis) if basis is not None else None,
            "nav_age_days": nav_age,
            "nav_is_published": nav_published,
            "nav_kind": nav_kind,
            "nav_est_error": lv.get("est_error"),
            "nav_staleness_days": lv.get("staleness_days"),
            "nav_rederived": nav_rederived,
            "nav_parser_status": parser_status,
            "nav_announcement_unparsed": unparsed,
            "nav_parse_failed_recorded": recorded_fail,
            "nav_parse_reason": parse_reason,
            "nav_parse_rate": parse_rate,
            "nav_announcements_held": ann_count,
            "last_parsed_nav_date": last_parsed,
            # units / sanity
            "discount": None if disc is None else round(disc, 6),
            "discount_stored": lv.get("discount_est"),
            "price_nav_ratio": diag["price_nav_ratio"],
            "unit_check_status": diag["unit_check_status"],
            "unit_check_reason": diag["unit_check_reason"],
            "suspected_scale_factor": diag["suspected_scale_factor"],
            "extreme_discount_flag": diag["extreme_discount_flag"],
            "price_vs_panel_ratio": None if pv is None else round(pv, 4),
            "price_history_check": price_hist_check,
            # signal inputs
            "disc_mu_36m": mu,
            "disc_sigma_36m": sd,
            "zscore_history_ok": zscore_ok,
            "zscore_min_months_required": zmin,
            "z_adj": lv.get("z_adj"),
            "alert_eligible": bool(lv.get("alert_eligible", False)),
            "in_live_table": bool(lv),
        })
    return pd.DataFrame(rows)


def classify_coverage(df: pd.DataFrame, params: dict) -> pd.DataFrame:
    """GREEN / AMBER / RED / EXCLUDED, decided in code and nowhere else.

    Read top to bottom: the first rule that fires wins, so the reported
    blocking issue is the one that would have to be fixed first.
    """
    thr = thresholds(params)
    p_fresh = int(thr["price"]["fresh_days"])
    p_stale = int(thr["price"]["stale_days"])
    n_fresh = int(thr["nav"]["fresh_days"])
    n_amber = int(thr["nav"]["amber_days"])

    status, reason, blocking, fix = [], [], [], []
    usable_price, usable_nav, valid_disc, sig_ready = [], [], [], []

    for r in df.to_dict("records"):
        if not r["monitoring_eligible"]:
            status.append(EXCLUDED)
            reason.append(r["exclusion_reason"] or "not_in_monitoring_universe")
            blocking.append("")
            fix.append("")
            usable_price.append(False); usable_nav.append(False)
            valid_disc.append(False); sig_ready.append(False)
            continue

        pa, na = r["price_age_days"], r["nav_age_days"]
        has_price = r["price"] is not None and pd.notna(r["price"])
        has_nav = r["nav"] is not None and pd.notna(r["nav"])
        up = bool(has_price and not r["price_is_fallback_panel"]
                  and pa is not None and 0 <= pa <= p_stale)
        un = bool(has_nav and na is not None and 0 <= na <= n_amber)
        units_ok = r["unit_check_status"] in ("ok", "extreme")
        vd = bool(up and un and units_ok and r["discount"] is not None)

        issues = []           # RED, in priority order
        # A date after the audit date is not evidence of anything. Letting a
        # negative age through would make a future-stamped observation the
        # freshest data we hold - look-ahead arriving as a bad timestamp.
        if pa is not None and pa < 0:
            issues.append("future_dated_price")
        if na is not None and na < 0:
            issues.append("future_dated_nav")
        if not has_price and not r["in_live_table"]:
            issues.append("not_in_live_table")
        elif not has_price and not r["ticker"]:
            issues.append("ticker_unresolved")
        elif not has_price:
            issues.append("no_price")
        elif r["price_is_fallback_panel"]:
            issues.append("stale_panel_price_only")
        elif pa is None or pa > p_stale:
            issues.append("stale_price")
        if not has_nav:
            issues.append("nav_announcement_unparsed"
                          if r["nav_announcement_unparsed"] else "no_nav")
        elif na is None or na > n_amber:
            issues.append("nav_too_stale")
        if r["unit_check_status"] == "suspect_scale":
            issues.append("suspected_unit_mismatch")

        if issues:
            status.append(RED)
            reason.append("; ".join(issues))
            blocking.append(issues[0])
            fix.append(FIXES.get(issues[0], ""))
            usable_price.append(up); usable_nav.append(un)
            valid_disc.append(vd); sig_ready.append(False)
            continue

        quals = []            # AMBER, in priority order
        if not r["zscore_history_ok"]:
            quals.append("insufficient_zscore_history")
        if r["unit_check_status"] in ("extreme", "metadata_conflict"):
            quals.append("extreme_discount_premium")
        if na is not None and na > n_fresh:
            quals.append("stale_nav")
        if r["nav_basis"] == 1:
            quals.append("rolled_forward_nav")
        if not r["nav_is_published"]:
            quals.append("monthly_report_nav_not_announcement"
                         if r["market"] == "ASX" else "aggregator_nav_only")
        if pa is not None and pa > p_fresh:
            quals.append("stale_price")

        if quals:
            status.append(AMBER)
            reason.append("; ".join(quals))
            blocking.append(quals[0])
            fix.append(FIXES.get(quals[0], ""))
            usable_price.append(up); usable_nav.append(un)
            valid_disc.append(vd); sig_ready.append(False)
            continue

        status.append(GREEN)
        reason.append("fresh price, published NAV, units consistent, "
                      "own discount history sufficient")
        blocking.append("")
        fix.append("")
        usable_price.append(up); usable_nav.append(un)
        valid_disc.append(vd); sig_ready.append(True)

    out = df.copy()
    out["usable_price"] = usable_price
    out["usable_nav"] = usable_nav
    out["valid_discount"] = valid_disc
    out["signal_ready"] = sig_ready
    out["coverage_status"] = status
    out["coverage_reason"] = reason
    out["blocking_issue"] = blocking
    out["recommended_fix"] = fix
    return out


# ------------------------------------------------------------------ summaries

def _pct(n, d):
    return None if not d else round(100.0 * n / d, 1)


def summarise(rows: pd.DataFrame, params: dict) -> dict:
    thr = thresholds(params)
    out = {}
    groups = {"UK": rows[rows["market"] == "UK"],
              "ASX": rows[rows["market"] == "ASX"],
              "COMBINED": rows}
    for label, g in groups.items():
        mon = g[g["monitoring_eligible"]]
        d = len(mon)
        fresh_price = int(mon["price_is_fresh"].sum())
        fallback = int(mon["price_is_fallback_panel"].sum())
        no_price = int(mon["price"].isna().sum())
        stale_price = d - fresh_price - fallback - no_price
        pub_nav = int((mon["nav_is_published"]
                       & (mon["nav_age_days"] <= thr["nav"]["fresh_days"])).sum())
        rolled = int((mon["nav_basis"] == 1).sum())
        no_nav = int(mon["nav"].isna().sum())
        stale_nav = int((mon["nav_age_days"] > thr["nav"]["fresh_days"]).sum())
        unparsed_ann = int(mon["nav_announcement_unparsed"].sum())
        parse_fail = int(mon["nav_parse_failed_recorded"].sum())
        out[label] = {
            "universe": {
                "registry_total": int(len(g)),
                "registry_labelled_live": int((g["aggregator_status"] == "live").sum()),
                "liveness_adjusted_live":
                    int(g["live_status"].isin(liveness.LIVE_STATUSES).sum()),
                "research_eligible": int(g["research_eligible"].sum()),
                "monitoring_eligible": d,
                "excluded": int(len(g) - d),
            },
            "prices": {
                "fresh": fresh_price, "fresh_pct": _pct(fresh_price, d),
                "stale": max(0, stale_price), "stale_pct": _pct(max(0, stale_price), d),
                "fallback_historical_panel": fallback,
                "fallback_pct": _pct(fallback, d),
                "no_price": no_price, "no_price_pct": _pct(no_price, d),
                "unresolved_ticker": int((mon["ticker_status"] == "unresolved").sum()),
            },
            "nav": {
                "fresh_published": pub_nav, "fresh_published_pct": _pct(pub_nav, d),
                "modelled_rolled_forward": rolled,
                "modelled_rolled_forward_pct": _pct(rolled, d),
                "stale": stale_nav, "stale_pct": _pct(stale_nav, d),
                "no_usable_nav": no_nav, "no_usable_nav_pct": _pct(no_nav, d),
                "announcement_held_but_unparsed": unparsed_ann,
                "announcement_held_but_unparsed_pct": _pct(unparsed_ann, d),
                "parser_failure_recorded": parse_fail,
                "parser_failure_recorded_pct": _pct(parse_fail, d),
            },
            "signal": {
                "usable_price": int(mon["usable_price"].sum()),
                "usable_price_pct": _pct(int(mon["usable_price"].sum()), d),
                "usable_nav": int(mon["usable_nav"].sum()),
                "usable_nav_pct": _pct(int(mon["usable_nav"].sum()), d),
                "normalised_units": int(mon["unit_check_status"]
                                        .isin(["ok", "extreme"]).sum()),
                "valid_discount": int(mon["valid_discount"].sum()),
                "valid_discount_pct": _pct(int(mon["valid_discount"].sum()), d),
                "zscore_history": int(mon["zscore_history_ok"].sum()),
                "zscore_history_pct": _pct(int(mon["zscore_history_ok"].sum()), d),
                "signal_ready": int(mon["signal_ready"].sum()),
                "signal_ready_pct": _pct(int(mon["signal_ready"].sum()), d),
            },
            "status": {s: int((mon["coverage_status"] == s).sum())
                       for s in (GREEN, AMBER, RED)},
            "excluded_by_reason": (g[~g["monitoring_eligible"]]["exclusion_reason"]
                                   .replace("", "unstated").value_counts().to_dict()),
        }
    return out


ISSUE_LABEL = {
    "not_in_live_table": "Live but absent from the live table",
    "future_dated_price": "Price dated after the audit date",
    "future_dated_nav": "NAV dated after the audit date",
    "stale_nav": "Somewhat stale NAV",
    "ticker_unresolved": "Ticker unresolved",
    "no_price": "No price returned",
    "stale_panel_price_only": "Stale historical panel price only",
    "stale_price": "Stale price",
    "no_nav": "No usable NAV",
    "nav_announcement_unparsed": "NAV announcement found but unparsed",
    "nav_too_stale": "NAV older than the amber window",
    "suspected_unit_mismatch": "Suspected unit mismatch",
    "insufficient_zscore_history": "Insufficient z-score history",
    "extreme_discount_premium": "Extreme discount/premium",
    "rolled_forward_nav": "Rolled-forward (modelled) NAV",
    "aggregator_nav_only": "Aggregator NAV only (no own route)",
    "monthly_report_nav_not_announcement": "ASX monthly report NTA (no announcement route)",
}


def failure_table(rows: pd.DataFrame) -> pd.DataFrame:
    """Every issue on every monitored fund, ranked - not just the first one."""
    mon = rows[rows["monitoring_eligible"]]
    recs = []
    for r in mon.to_dict("records"):
        if r["coverage_status"] == GREEN:
            continue
        for issue in [x for x in str(r["coverage_reason"]).split("; ") if x]:
            recs.append({"issue": ISSUE_LABEL.get(issue, issue), "issue_key": issue,
                         "severity": r["coverage_status"],
                         "market": r["market"], "ticker": r["ticker"],
                         "name": r["name"], "security_id": r["security_id"],
                         "nav_source": r["nav_source"],
                         "price_source": r["price_source"],
                         "price_age_days": r["price_age_days"],
                         "nav_age_days": r["nav_age_days"],
                         "unit_check_reason": r["unit_check_reason"],
                         "recommended_fix": FIXES.get(issue, "")})
    return pd.DataFrame(recs, columns=[
        "issue", "issue_key", "severity", "market", "ticker", "name",
        "security_id", "nav_source", "price_source", "price_age_days",
        "nav_age_days", "unit_check_reason", "recommended_fix"])


def failure_ranking(failures: pd.DataFrame) -> pd.DataFrame:
    if not len(failures):
        return pd.DataFrame(columns=["issue", "UK", "ASX", "Total"])
    piv = (failures.pivot_table(index="issue", columns="market",
                                values="security_id", aggfunc="count")
           .fillna(0).astype(int))
    for m in ("UK", "ASX"):
        if m not in piv.columns:
            piv[m] = 0
    piv["Total"] = piv[["UK", "ASX"]].sum(axis=1)
    return piv[["UK", "ASX", "Total"]].sort_values("Total", ascending=False) \
        .reset_index()


# ------------------------------------------------------------------ outputs

_PCT_COLS = {"nav_parse_rate", "discount", "discount_stored"}
_DATE_COLS = {"price_date", "nav_effective_date", "nav_announcement_date",
              "last_parsed_nav_date"}


def _format_sheet(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = ws.dimensions
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    for c in ws[1]:
        c.font = head
        c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for i, col in enumerate(df.columns, start=1):
        letter = get_column_letter(i)
        width = max([len(str(col))] +
                    [len(str(v)) for v in df[col].head(200).tolist()] or [10])
        ws.column_dimensions[letter].width = min(46, max(10, width + 2))
        fmt = None
        if col in _PCT_COLS:
            fmt = "0.0%"
        elif col in _DATE_COLS:
            fmt = "yyyy-mm-dd"
        elif col in ("price", "nav", "price_raw", "nav_stored_raw",
                     "price_nav_ratio", "disc_mu_36m", "disc_sigma_36m", "z_adj"):
            fmt = "0.0000"
        if fmt:
            for cell in ws[letter][1:]:
                cell.number_format = fmt
    # colour the verdict column so a scan reads at a glance
    if "coverage_status" in list(df.columns):
        idx = list(df.columns).index("coverage_status") + 1
        colours = {GREEN: "C6EFCE", AMBER: "FFEB9C", RED: "FFC7CE",
                   EXCLUDED: "D9D9D9"}
        for cell in ws[get_column_letter(idx)][1:]:
            if cell.value in colours:
                cell.fill = PatternFill("solid", fgColor=colours[cell.value])


def write_outputs(rows: pd.DataFrame, summary: dict, failures: pd.DataFrame,
                  ranking: pd.DataFrame, out_dir: Path = OUT_DIR,
                  meta: dict | None = None) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = {}

    rows.to_csv(out_dir / "coverage_audit.csv", index=False)
    paths["csv"] = str(out_dir / "coverage_audit.csv")
    failures.to_csv(out_dir / "coverage_failures.csv", index=False)
    paths["failures"] = str(out_dir / "coverage_failures.csv")

    mon = rows[rows["monitoring_eligible"]]
    tabs = {
        "Summary": _summary_frame(summary),
        "All Funds": mon,
        "UK": mon[mon["market"] == "UK"],
        "ASX": mon[mon["market"] == "ASX"],
        "Red Issues": mon[mon["coverage_status"] == RED],
        "Parser Failures": mon[mon["nav_announcement_unparsed"]].sort_values(
            ["nav_parse_failed_recorded", "market", "name"], ascending=[False, True, True]),
        "Ticker Issues": mon[(mon["ticker_status"] == "unresolved")
                             | mon["price_error"].notna()],
        "Unit Warnings": mon[mon["unit_check_status"].isin(
            ["suspect_scale", "extreme", "metadata_conflict"])],
        "Excluded": rows[~rows["monitoring_eligible"]],
        "Failure Ranking": ranking,
        "All Failures": failures,
    }
    xlsx = out_dir / "coverage_audit.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xl:
        for name, frame in tabs.items():
            frame = frame if len(frame) else frame.head(0)
            frame.to_excel(xl, sheet_name=name[:31], index=False)
            _format_sheet(xl.book[name[:31]], frame)
    paths["xlsx"] = str(xlsx)

    md = render_markdown(summary, ranking, rows, meta or {})
    (out_dir / "coverage_summary.md").write_text(md)
    paths["md"] = str(out_dir / "coverage_summary.md")

    (out_dir / "coverage_summary.json").write_text(
        json.dumps({"summary": summary, "meta": meta or {}}, indent=2, default=str))
    paths["json"] = str(out_dir / "coverage_summary.json")
    return paths


def _summary_frame(summary: dict) -> pd.DataFrame:
    recs = []
    for market, blocks in summary.items():
        for block, vals in blocks.items():
            if not isinstance(vals, dict):
                continue
            for k, v in vals.items():
                recs.append({"market": market, "section": block,
                             "metric": k, "value": v})
    return pd.DataFrame(recs)


def render_markdown(summary: dict, ranking: pd.DataFrame, rows: pd.DataFrame,
                    meta: dict) -> str:
    L = ["# Live fund coverage audit", "",
         f"Generated {meta.get('generated_at', '')} (as at "
         f"{meta.get('as_of', '')}).", "",
         "Run on demand only - this report has no schedule.", ""]
    if meta.get("refresh"):
        L += [f"Data refresh: `{meta['refresh']}`", ""]
    L += ["## Denominators", "",
          "| Denominator | UK | ASX | Combined |", "|---|--:|--:|--:|"]
    keys = [("registry_total", "Registry universe (all vehicles ever listed)"),
            ("registry_labelled_live", "Registry-labelled live (aggregator)"),
            ("liveness_adjusted_live", "Liveness-adjusted live (own filings)"),
            ("research_eligible", "Research eligible"),
            ("monitoring_eligible", "Monitoring eligible (audit denominator)"),
            ("excluded", "Excluded")]
    for k, label in keys:
        L.append(f"| {label} | {summary['UK']['universe'][k]} | "
                 f"{summary['ASX']['universe'][k]} | "
                 f"{summary['COMBINED']['universe'][k]} |")
    L.append("")

    for market in ("UK", "ASX", "COMBINED"):
        s = summary[market]
        d = s["universe"]["monitoring_eligible"]
        L += [f"## {market}", "", "```",
              f"Monitoring-eligible funds:  {d:>6}",
              f"Fresh price:                {s['prices']['fresh']:>6} / "
              f"{s['prices']['fresh_pct']}%",
              f"Usable price:               {s['signal']['usable_price']:>6} / "
              f"{s['signal']['usable_price_pct']}%",
              f"Usable NAV:                 {s['signal']['usable_nav']:>6} / "
              f"{s['signal']['usable_nav_pct']}%",
              f"Valid current discount:     {s['signal']['valid_discount']:>6} / "
              f"{s['signal']['valid_discount_pct']}%",
              f"Valid z-score history:      {s['signal']['zscore_history']:>6} / "
              f"{s['signal']['zscore_history_pct']}%",
              f"Fully signal-ready:         {s['signal']['signal_ready']:>6} / "
              f"{s['signal']['signal_ready_pct']}%",
              "",
              f"GREEN {s['status'][GREEN]}   AMBER {s['status'][AMBER]}   "
              f"RED {s['status'][RED]}",
              "```", "",
              "| Price | n | % |", "|---|--:|--:|",
              f"| Fresh | {s['prices']['fresh']} | {s['prices']['fresh_pct']}% |",
              f"| Stale | {s['prices']['stale']} | {s['prices']['stale_pct']}% |",
              f"| Historical panel fallback | {s['prices']['fallback_historical_panel']} "
              f"| {s['prices']['fallback_pct']}% |",
              f"| No price | {s['prices']['no_price']} | {s['prices']['no_price_pct']}% |",
              f"| Unresolved ticker | {s['prices']['unresolved_ticker']} | |", "",
              "NAV rows below are independent flags, not a partition: a NAV "
              "can be both modelled and stale, and a fund with a usable "
              "monthly NAV can still have an unparsed announcement.", "",
              "| NAV | n | % |", "|---|--:|--:|",
              f"| Fresh directly published | {s['nav']['fresh_published']} | "
              f"{s['nav']['fresh_published_pct']}% |",
              f"| Modelled / rolled forward | {s['nav']['modelled_rolled_forward']} | "
              f"{s['nav']['modelled_rolled_forward_pct']}% |",
              f"| Stale | {s['nav']['stale']} | {s['nav']['stale_pct']}% |",
              f"| No usable NAV | {s['nav']['no_usable_nav']} | "
              f"{s['nav']['no_usable_nav_pct']}% |",
              f"| NAV announcement held but no value parsed | "
              f"{s['nav']['announcement_held_but_unparsed']} | "
              f"{s['nav']['announcement_held_but_unparsed_pct']}% |",
              f"| ...of which the parser ran and failed | "
              f"{s['nav']['parser_failure_recorded']} | "
              f"{s['nav']['parser_failure_recorded_pct']}% |", ""]

    L += ["## Why coverage is missing", "",
          "| Issue | UK | ASX | Total |", "|---|--:|--:|--:|"]
    for r in ranking.to_dict("records"):
        L.append(f"| {r['issue']} | {r['UK']} | {r['ASX']} | {r['Total']} |")
    L.append("")

    mon = rows[rows["monitoring_eligible"]]
    red = mon[mon["coverage_status"] == RED]
    if len(red):
        L += ["## RED funds", "",
              "| Market | Ticker | Fund | Blocking issue | Recommended fix |",
              "|---|---|---|---|---|"]
        for r in red.sort_values(["market", "name"]).to_dict("records"):
            L.append(f"| {r['market']} | {r['ticker'] or ''} | {r['name']} | "
                     f"{ISSUE_LABEL.get(r['blocking_issue'], r['blocking_issue'])} | "
                     f"{r['recommended_fix']} |")
        L.append("")

    excl = rows[~rows["monitoring_eligible"]]
    L += ["## Excluded vehicles", "",
          f"{len(excl)} vehicles are outside the monitoring universe. They keep "
          "their rows in the Excluded tab with the reason; none is deleted.", "",
          "| Reason | n |", "|---|--:|"]
    for k, v in (excl["exclusion_reason"].replace("", "unstated")
                 .value_counts().head(20).items()):
        L.append(f"| {k} | {v} |")
    L += ["", "## How to read the verdicts", "",
          "- **GREEN** - fresh, credible price and a published NAV in "
          "consistent units, with enough of the fund's own discount history "
          "to z-score. A live signal can be produced.",
          "- **AMBER** - monitorable with a stated qualification: a "
          "rolled-forward or somewhat stale NAV, a somewhat stale price, an "
          "extreme but not impossible discount, or too little history to "
          "z-score.",
          "- **RED** - no reliable live signal is possible: no current price, "
          "an unresolved ticker, only a historical panel price, no usable "
          "NAV, an unparsed NAV announcement, or a suspected unit mismatch.",
          "- **EXCLUDED** - not part of the intended monitoring universe "
          "(research-policy exclusions, or not currently live).", ""]
    return "\n".join(L)


# ------------------------------------------------------------------ refresh

def refresh_data(markets: list[str]) -> dict:
    """Run the EXISTING nightly refresh, then report what it managed.

    No fetching logic lives here: this calls cli.nightly, which is the
    pipeline's own price and NAV refresh. A provider that fails is recorded
    and the audit continues over whatever was already on disk - a broken
    feed must degrade the report, not cancel it.
    """
    from . import cli
    started = datetime.now(timezone.utc)
    out = {"attempted_markets": markets, "started_at": started.isoformat(timespec="seconds")}
    try:
        rc = cli.nightly(markets)
        out["status"] = "ok" if rc == 0 else "completed_with_errors"
        out["exit_code"] = rc
    except Exception as exc:  # noqa: BLE001
        out["status"] = "failed"
        out["error"] = f"{type(exc).__name__}: {exc}"
        print(f"refresh FAILED ({out['error']}); auditing stored data instead")
    out["finished_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    return out


# ------------------------------------------------------------------ console

def print_console(summary: dict, ranking: pd.DataFrame, paths: dict) -> None:
    print("\nCEF LIVE COVERAGE AUDIT")
    print("=======================")
    for market in ("UK", "ASX"):
        s = summary[market]
        d = s["universe"]["monitoring_eligible"]
        print(f"\n{market}")
        print(f"  Live monitoring universe: {d}")
        print(f"  Signal ready: {s['signal']['signal_ready']} "
              f"({s['signal']['signal_ready_pct']}%)")
        print(f"  Amber: {s['status'][AMBER]}")
        print(f"  Red:   {s['status'][RED]}")
    print("\nTop blocking issues:")
    for i, r in enumerate(ranking.head(5).to_dict("records"), start=1):
        print(f"  {i}. {r['issue']}: {r['Total']}  (UK {r['UK']}, ASX {r['ASX']})")
    print("\nReport written to:")
    for k in ("md", "csv", "xlsx", "failures"):
        if paths.get(k):
            print(f"  {paths[k]}")


# ------------------------------------------------------------------ entry

def run(refresh: bool = False, markets: list[str] | None = None,
        out_dir: Path = OUT_DIR, as_of: date | None = None) -> dict:
    params = load_params()
    meta = {"generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "as_of": (as_of or datetime.now(timezone.utc).date()).isoformat(),
            "thresholds": thresholds(params)}
    if refresh:
        meta["refresh"] = refresh_data(markets or ["au", "uk"])

    reg = load_registry(params, as_of=as_of)
    live = load_live_table()
    meta["live_table_rows"] = int(len(live))
    meta["registry_rows"] = int(len(reg))
    if not len(live):
        print("WARNING: data/nta_live/latest.parquet is absent - every fund "
              "will audit as having no price and no NAV, which is a true "
              "statement about the data on disk, not a fund-level failure.")

    rows = classify_coverage(build_rows(reg, live, params, today=as_of), params)
    summary = summarise(rows, params)
    failures = failure_table(rows)
    ranking = failure_ranking(failures)
    paths = write_outputs(rows, summary, failures, ranking, out_dir, meta)
    print_console(summary, ranking, paths)
    return {"summary": summary, "paths": paths, "rows": rows,
            "failures": failures, "ranking": ranking, "meta": meta}


def email_report(out_dir: Path = OUT_DIR) -> bool:
    """Email the report that is already on disk, via the existing channel.

    Optional and never required: an unconfigured mailbox prints the summary
    and returns False rather than failing the run (see notify.notify).
    """
    from .notify import notify
    md = out_dir / "coverage_summary.md"
    if not md.exists():
        print(f"nothing to email: {md} does not exist")
        return False
    body = md.read_text()
    attach = [str(out_dir / f) for f in ("coverage_audit.xlsx",
                                         "coverage_audit.csv",
                                         "coverage_failures.csv")
              if (out_dir / f).exists()]
    return notify("live fund coverage audit", body[:60000], attachments=attach)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        prog="python -m cef_live.coverage_audit",
        description="On-demand coverage audit of the live fund universe. "
                    "Manual only - never scheduled.")
    ap.add_argument("--refresh", action="store_true",
                    help="run the existing price/NAV refresh first, then audit")
    ap.add_argument("--markets", default="au,uk",
                    help="markets to refresh when --refresh is given")
    ap.add_argument("--out-dir", default=str(OUT_DIR))
    ap.add_argument("--as-of", default=None,
                    help="audit as at this date (YYYY-MM-DD); defaults to today")
    ap.add_argument("--email-only", action="store_true",
                    help="email the report already in --out-dir; audit nothing")
    args = ap.parse_args(argv)
    if args.email_only:
        print("emailed:", email_report(Path(args.out_dir)))
        return 0
    as_of = (pd.Timestamp(args.as_of).date() if args.as_of else None)
    run(refresh=args.refresh,
        markets=[m.strip() for m in args.markets.split(",") if m.strip()],
        out_dir=Path(args.out_dir), as_of=as_of)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
